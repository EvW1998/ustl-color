import pandas as pd
import numpy as np
from openpyxl import Workbook
import math

LIST1 = ["R'", "G'", "B'", "", "R", "G", "B", "", "X", "Y", "Z", "", "L*", "a*", "b*", "C_ab*", "h_ab"]
color_data = pd.read_excel("../q1/tsv_d65.xlsx", sheet_name="Sheet")
white_point = color_data.iloc[:1, :3].to_numpy()[0]
M_RGB_inverse = np.array([[0.412453, 0.357580, 0.180423],
                          [0.212671, 0.715160, 0.072169],
                          [0.019334, 0.119193, 0.950227]])
a0 = 0.00304
r = 1 / 2.4

s = r / (a0 * (r - 1) + pow(a0, 1 - r))
d = 1 / (pow(a0, r) * (r - 1) + 1) - 1


def nonlinear_2_linear(t):
    if t <= a0 * s:
        return t / s
    else:
        return pow((t + d) / (1 + d), 1 / r)


nonlinear_2_linear_vec = np.vectorize(nonlinear_2_linear)


def xyz_2_cielab_transfer(t):
    if t > (24 / 116) ** 3:
        return t ** (1 / 3)

    return 841 / 108 * t + 16 / 116


sRGB_data = pd.read_excel("../q5/camera_sRGB.xlsx", sheet_name="Sheet")
sRGB = sRGB_data.iloc[:, 1:].to_numpy() / 255
rgb = nonlinear_2_linear_vec(sRGB)
xyz = np.matmul(M_RGB_inverse, rgb.T).T * 100


wb = Workbook()
ws = wb.active

for i in range(len(LIST1)):
    ws.cell(1, i + 2, LIST1[i])

for i in range(24):
    ws.cell(i + 2, 1, i + 1)
    ws.cell(i + 2, 2, sRGB[i][0])
    ws.cell(i + 2, 3, sRGB[i][1])
    ws.cell(i + 2, 4, sRGB[i][2])
    ws.cell(i + 2, 6, rgb[i][0])
    ws.cell(i + 2, 7, rgb[i][1])
    ws.cell(i + 2, 8, rgb[i][2])
    ws.cell(i + 2, 10, xyz[i][0])
    ws.cell(i + 2, 11, xyz[i][1])
    ws.cell(i + 2, 12, xyz[i][2])

    x_xw = xyz_2_cielab_transfer(xyz[i][0] / white_point[0])
    y_yw = xyz_2_cielab_transfer(xyz[i][1] / white_point[1])
    z_zw = xyz_2_cielab_transfer(xyz[i][2] / white_point[2])

    a = 500 * (x_xw - y_yw)
    b = 200 * (y_yw - z_zw)

    ws.cell(i + 2, 14, 116 * y_yw - 16)
    ws.cell(i + 2, 15, a)
    ws.cell(i + 2, 16, b)
    ws.cell(i + 2, 17, (a * a + b * b) ** 0.5)
    ws.cell(i + 2, 18, math.atan(b / a))

wb.save("sRGB_2_Lab.xlsx")
