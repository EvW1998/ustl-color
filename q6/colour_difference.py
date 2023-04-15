import pandas as pd
import numpy as np
from openpyxl import Workbook
import math


LIST1 = ["delta E by Lab", "", "delta E by LCH"]

camera_data = pd.read_excel("sRGB_2_Lab.xlsx", sheet_name="Sheet")
camera = camera_data.iloc[:, 13:].to_numpy()
color_card_data = pd.read_excel("../q3/xyz_2_CIELAB.xlsx", sheet_name="Sheet")
color_card = color_card_data.iloc[:, 5:].to_numpy()

delta_e_lab = pow(camera[:, 0:3] - color_card[:, 0:3], 2)
delta_e_lab = np.sqrt(np.sum(delta_e_lab, axis=1))

delta_e_lch = pow(camera[:, 0] - color_card[:, 0], 2) + pow(camera[:, 3] - color_card[:, 3], 2)
delta_H = 2 * np.sqrt(camera[:, 3] * color_card[:, 3]) * np.sin((camera[:, 4] - color_card[:, 4]) / 2)
delta_e_lch += pow(delta_H, 2)
delta_e_lch = np.sqrt(delta_e_lch)


wb = Workbook()
ws = wb.active

for i in range(len(LIST1)):
    ws.cell(1, i + 2, LIST1[i])

for i in range(24):
    ws.cell(i + 2, 1, i + 1)
    ws.cell(i + 2, 2, delta_e_lab[i])

    ws.cell(i + 2, 4, delta_e_lch[i])

wb.save("delta_e.xlsx")
