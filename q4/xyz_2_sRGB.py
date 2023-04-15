import pandas as pd
import numpy as np
from openpyxl import Workbook

LIST1 = ["X", "Y", "Z", "", "R", "G", "B", "", "R'", "G'", "B'"]
M_RGB = np.array([[3.240479, -1.537150, -0.498535],
                  [-0.969256, 1.875992, 0.041556],
                  [0.055648, -0.204043, 1.057311]])
a_0 = 0.00304
r = 1 / 2.4

s = r / (a_0 * (r - 1) + pow(a_0, 1 - r))
d = 1 / (pow(a_0, r) * (r - 1) + 1) - 1


def gamma_correction(a):
    if a <= a_0:
        return s * a
    else:
        return (1 + d) * pow(a, r) - d


gamma_correction_vec = np.vectorize(gamma_correction)


def range_correction(a):
    if a < 0:
        return 0
    elif a > 1:
        return 1

    return a


range_correction_vec = np.vectorize(range_correction)

color_data = pd.read_excel("../q1/tsv_d65.xlsx", sheet_name="Sheet")
color_point = color_data.iloc[:, 5:8].to_numpy() / 100
rgb = range_correction_vec(np.matmul(M_RGB, color_point.T).T)
# rgb = np.matmul(M_RGB, color_point.T).T
srgb = gamma_correction_vec(rgb)

wb = Workbook()
ws = wb.active

for i in range(len(LIST1)):
    ws.cell(1, i + 2, LIST1[i])

for i in range(len(color_point)):
    ws.cell(i + 2, 1, i + 1)
    ws.cell(i + 2, 2, color_point[i][0])
    ws.cell(i + 2, 3, color_point[i][1])
    ws.cell(i + 2, 4, color_point[i][2])
    ws.cell(i + 2, 6, rgb[i][0])
    ws.cell(i + 2, 7, rgb[i][1])
    ws.cell(i + 2, 8, rgb[i][2])
    ws.cell(i + 2, 10, srgb[i][0])
    ws.cell(i + 2, 11, srgb[i][1])
    ws.cell(i + 2, 12, srgb[i][2])
    ws.cell(i + 2, 14, srgb[i][0] * 255)
    ws.cell(i + 2, 15, srgb[i][1] * 255)
    ws.cell(i + 2, 16, srgb[i][2] * 255)

wb.save("xyz_2_sRGB.xlsx")
