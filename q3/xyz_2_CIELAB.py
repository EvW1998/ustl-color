import pandas as pd
from openpyxl import Workbook
import math


def xyz_2_cielab_transfer(t):
    if t > (24 / 116) ** 3:
        return t ** (1 / 3)

    return 841 / 108 * t + 16 / 116


LIST1 = ["X", "Y", "Z"]
LIST2 = ["L*", "a*", "b*", "C_ab*", "h_ab"]

color_data = pd.read_excel("../q1/tsv_d65.xlsx", sheet_name="Sheet")
white_point = color_data.iloc[:1, :3].to_numpy()[0]
color_point = color_data.iloc[:, 5:8].to_numpy()

wb = Workbook()
ws = wb.active

for i in range(len(LIST1)):
    ws.cell(1, i + 2, LIST1[i])

for i in range(len(LIST2)):
    ws.cell(1, i + 6, LIST2[i])

for i in range(len(color_point)):
    ws.cell(i + 2, 1, i + 1)
    ws.cell(i + 2, 2, color_point[i][0])
    ws.cell(i + 2, 3, color_point[i][1])
    ws.cell(i + 2, 4, color_point[i][2])

    x_xw = xyz_2_cielab_transfer(color_point[i][0] / white_point[0])
    y_yw = xyz_2_cielab_transfer(color_point[i][1] / white_point[1])
    z_zw = xyz_2_cielab_transfer(color_point[i][2] / white_point[2])

    a = 500 * (x_xw - y_yw)
    b = 200 * (y_yw - z_zw)

    ws.cell(i + 2, 6, 116 * y_yw - 16)
    ws.cell(i + 2, 7, a)
    ws.cell(i + 2, 8, b)
    ws.cell(i + 2, 9, (a * a + b * b) ** 0.5)
    ws.cell(i + 2, 10, math.atan(b / a))

wb.save("xyz_2_CIELAB.xlsx")
