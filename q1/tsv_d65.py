import pandas as pd
from openpyxl import Workbook

LIST1 = ["Xw", "Yw", "Zw"]
LIST2 = ["X", "Y", "Z"]
LIST3 = ["x", "y", "z"]

color_data = pd.read_excel("../data.xlsx", sheet_name="24Ref")
wt_data = pd.read_excel("../data.xlsx", sheet_name="WT")
color_24Ref = color_data.iloc[:, 1:].to_numpy().T
color_wt = wt_data.iloc[:, 2:5].to_numpy().T
# color_wt = wt_data.iloc[:, 7:10].to_numpy().T

sum_wt = sum(color_wt[0]) + sum(color_wt[1]) + sum(color_wt[2])

wb = Workbook()
ws = wb.active

for i in range(len(LIST1)):
    ws.cell(1, i + 1, LIST1[i])
    ws.cell(2, i + 1, sum(color_wt[i]))
    ws.cell(4, i + 1, sum(color_wt[i]) / sum_wt)
    ws.cell(1, i + 6, LIST2[i])
    ws.cell(1, i + 10, LIST3[i])

for i in range(len(color_24Ref)):
    ws.cell(i + 2, 5, i + 1)
    x, y, z = 0, 0, 0
    for j in range(len(color_24Ref[i])):
        x += color_24Ref[i][j] * color_wt[0][j]
        y += color_24Ref[i][j] * color_wt[1][j]
        z += color_24Ref[i][j] * color_wt[2][j]

    x = x / 100
    y = y / 100
    z = z / 100

    ws.cell(i + 2, 6, x)
    ws.cell(i + 2, 7, y)
    ws.cell(i + 2, 8, z)

    ws.cell(i + 2, 10, x / (x + y + z))
    ws.cell(i + 2, 11, y / (x + y + z))
    ws.cell(i + 2, 12, z / (x + y + z))

wb.save("tsv_d65.xlsx")
