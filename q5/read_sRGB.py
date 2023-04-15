from PIL import Image
import numpy as np
from openpyxl import Workbook


LIST1 = ["R", "G", "B"]


def find_average_srgb(input_file):
    im = Image.open(input_file, "r")
    pixel_values = list(im.getdata())
    return np.round(np.mean(np.array(pixel_values), axis=0))


wb = Workbook()
ws = wb.active

for i in range(len(LIST1)):
    ws.cell(1, i + 2, LIST1[i])

for i in range(24):
    rgb = find_average_srgb(str(i + 1) + ".jpg")
    print(rgb/255)

    ws.cell(i + 2, 1, i + 1)
    ws.cell(i + 2, 2, rgb[0])
    ws.cell(i + 2, 3, rgb[1])
    ws.cell(i + 2, 4, rgb[2])

wb.save("camera_sRGB.xlsx")
