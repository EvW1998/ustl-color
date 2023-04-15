import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import torch
from openpyxl import Workbook


seed = 39
torch.manual_seed(seed)


def xyz_2_cielab_transfer(t):
    if t > (24 / 116) ** 3:
        return t ** (1 / 3)

    return 841 / 108 * t + 16 / 116


def calculate_color_difference(input_xyz):
    predict_Lab = []

    for index in input_xyz:
        x_xw = xyz_2_cielab_transfer(index[0] / white_point[0])
        y_yw = xyz_2_cielab_transfer(index[1] / white_point[1])
        z_zw = xyz_2_cielab_transfer(index[2] / white_point[2])

        predict_Lab.append([116 * y_yw - 16, 500 * (x_xw - y_yw), 200 * (y_yw - z_zw)])

    predict_Lab = np.array(predict_Lab)
    delta_e_lab = pow(predict_Lab - card_Lab, 2)
    delta_e_lab = np.sqrt(np.sum(delta_e_lab, axis=1))

    return delta_e_lab


# 设置超参数
lr = 0.01  # 学习率
epochs = 15000  # 训练轮数
n_feature = 3  # 输入特征
n_hidden = 10  # 隐层节点数
n_output = 3  # 输出

# 1.准备数据
camera_sRGB_table = pd.read_excel("../q5/camera_sRGB.xlsx", sheet_name="Sheet")
camera_sRGB = camera_sRGB_table.iloc[:, 1:].to_numpy()/255
xyz_table = pd.read_excel("../q1/tsv_d65.xlsx", sheet_name="Sheet")
xyz = xyz_table.iloc[:, 5:8].to_numpy()
card_Lab_table = pd.read_excel("../q3/xyz_2_CIELAB.xlsx", sheet_name="Sheet")
card_Lab = card_Lab_table.iloc[:, 5:8].to_numpy()
color_data = pd.read_excel("../q1/tsv_d65.xlsx", sheet_name="Sheet")
white_point = color_data.iloc[:1, :3].to_numpy()[0]


x_train = camera_sRGB
y_train = xyz

x_test = camera_sRGB
y_test = xyz

# 将数据类型转换为tensor方便pytorch使用
x_train = torch.FloatTensor(x_train)
y_train = torch.FloatTensor(y_train)
x_test = torch.FloatTensor(x_test)
y_test = torch.FloatTensor(y_test)


# 2.定义BP神经网络
class BPNetModel(torch.nn.Module):
    def __init__(self, n_feature, n_hidden, n_output):
        super(BPNetModel, self).__init__()
        self.linear_relu1 = torch.nn.Linear(n_feature, n_hidden)
        self.linear5 = torch.nn.Linear(n_hidden, n_output)

        # self.linear_relu1 = torch.nn.Linear(n_feature, 56)
        # self.linear_relu2 = torch.nn.Linear(56, 128)
        # self.linear_relu3 = torch.nn.Linear(256, 256)
        # self.linear_relu4 = torch.nn.Linear(256, 256)
        # self.linear5 = torch.nn.Linear(128, n_output)

    def forward(self, x):
        out = self.linear_relu1(x)
        out = torch.nn.functional.relu(out)

        out = self.linear5(out)
        return out

        # out = self.linear_relu1(x)
        # out = torch.nn.functional.relu(out)
        #
        # out = self.linear_relu2(out)
        # out = torch.nn.functional.relu(out)

        # out = self.linear_relu3(out)
        # out = torch.nn.functional.relu(out)
        #
        # out = self.linear_relu4(out)
        # out = torch.nn.functional.relu(out)

        # out = self.linear5(out)
        # return out


# 3.定义优化器和损失函数
net = BPNetModel(n_feature=n_feature, n_hidden=n_hidden, n_output=n_output)  # 调用网络
optimizer = torch.optim.Adam(net.parameters(), lr=lr)  # 使用Adam优化器，并设置学习率
loss_fun = torch.nn.MSELoss(reduction='mean')

# 4.训练数据
loss_steps = np.zeros(epochs)  # 构造一个array([ 0., 0., 0., 0., 0.])里面有epochs个0
accuracy_steps = np.zeros(epochs)

for epoch in range(epochs):
    y_pred = net(x_train)  # 前向传播
    loss = loss_fun(y_pred, y_train)  # 预测值和真实值对比
    optimizer.zero_grad()  # 梯度清零
    loss.backward()  # 反向传播
    optimizer.step()  # 更新梯度
    loss_steps[epoch] = loss.item()  # 保存loss
    running_loss = loss.item()

    if epoch % 1000 == 0:
        print(f"第{epoch}次训练，loss={running_loss}".format(epoch, running_loss))

    with torch.no_grad():  # 下面是没有梯度的计算,主要是测试集使用，不需要再计算梯度了
        y_pred = net(x_test)
        color_difference = calculate_color_difference(y_pred.numpy())
        accuracy_steps[epoch] = np.mean(color_difference)

        if epoch % 1000 == 0:
            print("测试预测三刺激值的平均色差", accuracy_steps[epoch])

print("最终测试结果")
r = y_pred.numpy()
print(r)
print("测试预测三刺激值的平均色差", accuracy_steps[-1])
print(color_difference)


print(net.state_dict())

# 5.绘制损失函数和精度
fig_name = "Color_Correction_BPNet"
fontsize = 24
fig, (ax1, ax2) = plt.subplots(2, figsize=(15, 12), sharex=True)

plt.rcParams['text.usetex'] = True
font1 = {'family': 'serif', 'color': 'blue', 'size': 26}
font2 = {'family': 'serif', 'color': 'darkred', 'size': 20}

ax1.plot(accuracy_steps, linewidth=3)
ax1.set_ylabel(r"Colour Difference $\Delta E^{*}_{ab}$", fontdict=font2)
# ax1.set_title(fig_name, fontdict=font1)
ax2.plot(loss_steps, linewidth=3)
ax2.set_ylabel("Train lss", fontdict=font2)
ax2.set_xlabel("Epochs", fontdict=font2)
plt.tight_layout()
plt.savefig(fig_name + '.png')
# plt.show()

wb = Workbook()
ws = wb.active

for i in range(24):
    ws.cell(i + 1, 1, r[i][0])
    ws.cell(i + 1, 2, r[i][1])
    ws.cell(i + 1, 3, r[i][2])

    ws.cell(i + 1, 5, color_difference[i])

ws.cell(1, 7, accuracy_steps[-1])

wb.save("bp_network_model.xlsx")
