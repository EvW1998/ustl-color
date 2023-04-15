import pandas as pd
import numpy as np
from openpyxl import Workbook


def xyz_2_cielab_transfer(t):
    if t > (24 / 116) ** 3:
        return t ** (1 / 3)

    return 841 / 108 * t + 16 / 116


def get_j_combination(n, space):
    if space == 1:
        return [[n]]
    elif n == 0:
        return [[0] * space]
    else:
        result = []

        for i in range(n + 1):
            sub_result = get_j_combination(n - i, space - 1)

            for s in sub_result:
                result.append([i] + s)

        return result


def get_combination_n(n):
    result = []

    for i in range(n + 1):
        result += get_j_combination(i, 3)

    return np.array(result)


def generate_color_matrix(combination, rgb):
    matrix = np.ones((24, 1))
    
    for i in combination[1:]:
        vector = []

        for j in rgb:
            r = 1
            for k in range(3):
                r *= pow(j[k], i[k])

            vector.append(r)

        vector = np.array(vector)
        vector = np.reshape(vector, (1, vector.size)).T
        matrix = np.hstack((matrix, vector))

    return matrix


LIST1 = ["delta E", "", "average"]
LIST2 = ["a_x", "a_y", "a_z"]
camera_sRGB_table = pd.read_excel("../q5/camera_sRGB.xlsx", sheet_name="Sheet")
camera_sRGB = camera_sRGB_table.iloc[:, 1:].to_numpy()
xyz_table = pd.read_excel("../q1/tsv_d65.xlsx", sheet_name="Sheet")
xyz = xyz_table.iloc[:, 5:8].to_numpy()
card_Lab_table = pd.read_excel("../q3/xyz_2_CIELAB.xlsx", sheet_name="Sheet")
card_Lab = card_Lab_table.iloc[:, 5:8].to_numpy()
color_data = pd.read_excel("../q1/tsv_d65.xlsx", sheet_name="Sheet")
white_point = color_data.iloc[:1, :3].to_numpy()[0]
N = 3

j_combination = get_combination_n(N)
print(j_combination)
v = generate_color_matrix(j_combination, camera_sRGB)
A = np.matmul(np.matmul(np.linalg.inv(np.matmul(v.T, v)), v.T), xyz).T
print(A)

u = np.matmul(A, v.T).T
predict_Lab = []

for index in u:
    x_xw = xyz_2_cielab_transfer(index[0] / white_point[0])
    y_yw = xyz_2_cielab_transfer(index[1] / white_point[1])
    z_zw = xyz_2_cielab_transfer(index[2] / white_point[2])

    predict_Lab.append([116 * y_yw - 16, 500 * (x_xw - y_yw), 200 * (y_yw - z_zw)])

predict_Lab = np.array(predict_Lab)

delta_e_lab = pow(predict_Lab - card_Lab, 2)
delta_e_lab = np.sqrt(np.sum(delta_e_lab, axis=1))
print(delta_e_lab)
print("Average delta e: ", np.mean(delta_e_lab))


wb = Workbook()
ws = wb.active

for i in range(A.shape[0]):
    for j in range(A.shape[1]):
        ws.cell(j + 2, i + 1, A[i][j])

for i in range(len(LIST1)):
    ws.cell(1, A.shape[0] + 3 + i, LIST1[i])

for i in range(len(LIST2)):
    ws.cell(1, i + 1, LIST2[i])

for i in range(24):
    ws.cell(i + 2, A.shape[0] + 2, i + 1)
    ws.cell(i + 2, A.shape[0] + 3, delta_e_lab[i])

    ws.cell(i + 2, 10, u[i][0])
    ws.cell(i + 2, 11, u[i][1])
    ws.cell(i + 2, 12, u[i][2])


ws.cell(2, A.shape[0] + 5, np.mean(delta_e_lab))

wb.save(str(N) + "_th polynomial model.xlsx")
